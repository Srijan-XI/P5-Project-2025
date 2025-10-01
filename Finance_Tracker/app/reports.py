from flask import Blueprint, render_template, send_file, request, flash, redirect, url_for
from flask_login import login_required, current_user
from app.models import Transaction, Category, db
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, desc
from datetime import datetime, timedelta
import tempfile
import os

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/')
@login_required
def reports_dashboard():
    # Get date range from request args
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category_id = request.args.get('category')
    
    # Default to current month if no dates provided
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    
    # Build query
    query = Transaction.query.filter_by(user_id=current_user.id)
    
    if start_date:
        query = query.filter(Transaction.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Transaction.date <= datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    if category_id:
        query = query.filter_by(category_id=int(category_id))
    
    transactions = query.order_by(desc(Transaction.date)).all()
    
    # Calculate summary
    total_income = sum(t.amount for t in transactions if t.type == 'income')
    total_expenses = sum(t.amount for t in transactions if t.type == 'expense')
    net_balance = total_income - total_expenses
    
    # Category breakdown
    category_summary = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.type == 'expense'
    )
    
    if start_date:
        category_summary = category_summary.filter(Transaction.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        category_summary = category_summary.filter(Transaction.date <= datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    
    category_summary = category_summary.group_by(Category.name).order_by(desc('total')).all()
    
    categories = Category.query.all()
    
    return render_template('reports/dashboard.html',
                         transactions=transactions,
                         total_income=total_income,
                         total_expenses=total_expenses,
                         net_balance=net_balance,
                         category_summary=category_summary,
                         categories=categories,
                         start_date=start_date,
                         end_date=end_date,
                         selected_category=category_id)

@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    # Get filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category_id = request.args.get('category')
    
    # Build query
    query = Transaction.query.filter_by(user_id=current_user.id)
    
    if start_date:
        query = query.filter(Transaction.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Transaction.date <= datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    if category_id:
        query = query.filter_by(category_id=int(category_id))
    
    transactions = query.order_by(desc(Transaction.date)).all()
    
    # Create PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, f'Finance Report - {current_user.username}', ln=True, align='C')
    
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M")}', ln=True)
    
    if start_date and end_date:
        pdf.cell(0, 10, f'Period: {start_date} to {end_date}', ln=True)
    
    pdf.ln(10)
    
    # Summary
    total_income = sum(t.amount for t in transactions if t.type == 'income')
    total_expenses = sum(t.amount for t in transactions if t.type == 'expense')
    
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Summary:', ln=True)
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 8, f'Total Income: ₹{total_income:,.2f}', ln=True)
    pdf.cell(0, 8, f'Total Expenses: ₹{total_expenses:,.2f}', ln=True)
    pdf.cell(0, 8, f'Net Balance: ₹{total_income - total_expenses:,.2f}', ln=True)
    
    pdf.ln(10)
    
    # Transactions table
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(30, 8, 'Date', 1)
    pdf.cell(40, 8, 'Title', 1)
    pdf.cell(25, 8, 'Amount', 1)
    pdf.cell(25, 8, 'Type', 1)
    pdf.cell(30, 8, 'Category', 1)
    pdf.ln()
    
    pdf.set_font('Arial', '', 9)
    for transaction in transactions:
        pdf.cell(30, 8, transaction.date.strftime('%Y-%m-%d'), 1)
        pdf.cell(40, 8, transaction.title[:20], 1)
        pdf.cell(25, 8, f'₹{transaction.amount:,.2f}', 1)
        pdf.cell(25, 8, transaction.type.capitalize(), 1)
        category_name = transaction.category.name if transaction.category else 'N/A'
        pdf.cell(30, 8, category_name[:15], 1)
        pdf.ln()
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf.output(temp_file.name)
    
    return send_file(temp_file.name, 
                    as_attachment=True, 
                    download_name=f'finance_report_{datetime.now().strftime("%Y%m%d")}.pdf',
                    mimetype='application/pdf')

@reports_bp.route('/export/excel')
@login_required
def export_excel():
    # Get filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category_id = request.args.get('category')
    
    # Build query
    query = Transaction.query.filter_by(user_id=current_user.id)
    
    if start_date:
        query = query.filter(Transaction.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Transaction.date <= datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    if category_id:
        query = query.filter_by(category_id=int(category_id))
    
    transactions = query.order_by(desc(Transaction.date)).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Finance Report"
    
    # Header
    headers = ['Date', 'Title', 'Amount', 'Type', 'Category', 'Description']
    header_fill = PatternFill(start_color='366992', end_color='366992', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for row, transaction in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=transaction.date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=transaction.title)
        ws.cell(row=row, column=3, value=transaction.amount)
        ws.cell(row=row, column=4, value=transaction.type.capitalize())
        ws.cell(row=row, column=5, value=transaction.category.name if transaction.category else 'N/A')
        ws.cell(row=row, column=6, value=transaction.description or '')
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    total_income = sum(t.amount for t in transactions if t.type == 'income')
    total_expenses = sum(t.amount for t in transactions if t.type == 'expense')
    
    summary_ws['A1'] = 'Finance Summary'
    summary_ws['A1'].font = Font(bold=True, size=14)
    
    summary_ws['A3'] = 'Total Income:'
    summary_ws['B3'] = total_income
    summary_ws['A4'] = 'Total Expenses:'
    summary_ws['B4'] = total_expenses
    summary_ws['A5'] = 'Net Balance:'
    summary_ws['B5'] = total_income - total_expenses
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    return send_file(temp_file.name,
                    as_attachment=True,
                    download_name=f'finance_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')